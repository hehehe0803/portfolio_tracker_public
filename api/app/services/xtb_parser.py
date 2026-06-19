"""
XTB Statement Parser - Parse XTB XLSX statement files

Parses XTB statement files to extract:
- Open positions
- Closed positions  
- Cash operations (deposits, withdrawals, trades)

Returns normalized transaction objects with:
- id, date, type, amount, currency, symbol, description
"""

import hashlib
import re
import subprocess
import unicodedata
from abc import ABC, abstractmethod
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from email import policy
from email.parser import BytesParser
from enum import Enum
from html.parser import HTMLParser
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from pydantic import BaseModel, Field


class TransactionType(str, Enum):
    """Transaction types for XTB statements."""
    OPEN_POSITION = "OPEN_POSITION"
    CLOSE_POSITION = "CLOSE_POSITION"
    DEPOSIT = "DEPOSIT"
    WITHDRAWAL = "WITHDRAWAL"
    DIVIDEND = "DIVIDEND"
    COMMISSION = "COMMISSION"
    SWAP = "SWAP"
    STAMP_DUTY = "STAMP_DUTY"
    OTHER = "OTHER"


class PositionType(str, Enum):
    """Position side type."""
    BUY = "BUY"
    SELL = "SELL"


class XTBError(Exception):
    """Base exception for XTB parsing errors."""
    pass


class XTBFormatError(XTBError):
    """Raised when XTB file format is invalid or missing required columns."""
    pass


class XTBFileNotFoundError(XTBError):
    """Raised when XTB file does not exist."""
    pass


class XTBBaseModel(BaseModel):
    """Base model for XTB objects with common fields."""
    id: str = Field(..., description="Unique identifier")
    date: datetime = Field(..., description="Transaction date/time")
    description: str = Field(..., description="Transaction description/comment")
    
    class Config:
        json_encoders = {
            datetime: lambda d: d.isoformat(),
        }


@dataclass
class OpenPosition:
    """Represents an open position from XTB statement."""
    position_id: int
    symbol: str
    position_type: PositionType
    volume: Decimal
    open_time: datetime
    open_price: Decimal
    market_price: Optional[Decimal]
    purchase_value: Decimal
    commission: Decimal
    swap: Decimal
    rollover: Decimal
    gross_pl: Decimal
    comment: str
    
    def fingerprint(self) -> str:
        """Generate deterministic fingerprint for deduplication."""
        return generate_fingerprint(
            broker="XTB",
            tx_type=TransactionType.OPEN_POSITION,
            timestamp=self.open_time,
            amount=self.purchase_value,
            symbol=self.symbol
        )
    
    def to_transaction(self) -> 'XTBTransaction':
        """Convert to normalized transaction object."""
        return XTBTransaction(
            id=str(self.position_id),
            date=self.open_time,
            tx_type=TransactionType.OPEN_POSITION,
            amount=-self.purchase_value,
            currency=None,  # Determined from account
            symbol=self.symbol,
            description=self.comment or f"OPEN {self.position_type.value} {self.volume} {self.symbol}"
        )


@dataclass
class ClosedPosition:
    """Represents a closed position from XTB statement."""
    position_id: int
    symbol: str
    position_type: PositionType
    volume: Decimal
    open_time: datetime
    open_price: Decimal
    close_time: datetime
    close_price: Decimal
    purchase_value: Decimal
    sale_value: Decimal
    sl: Optional[Decimal]
    tp: Optional[Decimal]
    margin: Decimal
    commission: Decimal
    swap: Decimal
    rollover: Decimal
    gross_pl: Decimal
    comment: str
    
    def fingerprint(self) -> str:
        """Generate deterministic fingerprint for deduplication."""
        return generate_fingerprint(
            broker="XTB",
            tx_type=TransactionType.CLOSE_POSITION,
            timestamp=self.close_time,
            amount=self.sale_value,
            symbol=self.symbol
        )
    
    def to_transaction(self) -> 'XTBTransaction':
        """Convert to normalized transaction object."""
        # Net effect = sale value - purchase value (for closed positions)
        net_amount = self.sale_value - self.purchase_value
        return XTBTransaction(
            id=str(self.position_id),
            date=self.close_time,
            tx_type=TransactionType.CLOSE_POSITION,
            amount=net_amount,
            currency=None,  # Determined from account
            symbol=self.symbol,
            description=self.comment or f"CLOSE {self.position_type.value} {self.volume} {self.symbol}"
        )


@dataclass
class CashOperation:
    """Represents a cash operation (deposit, withdrawal, trade settlement) from XTB."""
    operation_id: int
    operation_type: TransactionType
    time: datetime
    comment: str
    symbol: Optional[str]
    amount: Decimal
    
    def fingerprint(self) -> str:
        """Generate deterministic fingerprint for deduplication."""
        return generate_fingerprint(
            broker="XTB",
            tx_type=self.operation_type,
            timestamp=self.time,
            amount=self.amount,
            symbol=self.symbol or ""
        )
    
    def to_transaction(self) -> 'XTBTransaction':
        """Convert to normalized transaction object."""
        return XTBTransaction(
            id=str(self.operation_id),
            date=self.time,
            tx_type=self.operation_type,
            amount=self.amount,
            currency=None,  # Determined from account
            symbol=self.symbol,
            description=self.comment
        )


@dataclass
class XTBDailyStatementTrade:
    """Represents one executed trade row from an XTB daily PDF statement."""
    order_id: int
    symbol: str
    instrument_name: str
    position_type: PositionType
    quantity: Decimal
    trade_time: datetime
    execution_price: Decimal
    total_value: Decimal
    asset_type: str
    currency: str
    fx_rate: Decimal
    conversion_fee: Decimal
    commission: Decimal
    total_cost: Decimal

    def to_transaction(self) -> 'XTBTransaction':
        """Convert daily executed trade into the existing transaction shape."""
        signed_amount = (
            -self.total_value
            if self.position_type == PositionType.BUY
            else self.total_value
        )
        description = (
            f"{self.position_type.value} {self.quantity} {self.symbol} "
            f"price={self.execution_price} value={self.total_value} "
            f"fx_rate={self.fx_rate} conversion_fee={self.conversion_fee} "
            f"commission={self.commission} total_cost={self.total_cost}"
        )
        return XTBTransaction(
            id=str(self.order_id),
            date=self.trade_time,
            tx_type=TransactionType.OPEN_POSITION
            if self.position_type == PositionType.BUY
            else TransactionType.CLOSE_POSITION,
            amount=signed_amount,
            currency=self.currency,
            symbol=self.symbol,
            description=description,
        )


class XTBTransaction(XTBBaseModel):
    """Normalized transaction object for XTB data."""
    tx_type: TransactionType = Field(..., alias="type", description="Transaction type")
    amount: Decimal = Field(..., description="Transaction amount (positive = inflow, negative = outflow)")
    currency: Optional[str] = Field(None, description="Currency code")
    symbol: Optional[str] = Field(None, description="Asset symbol/ticker")
    
    class Config:
        populate_by_name = True
        json_encoders = {
            datetime: lambda d: d.isoformat(),
            Decimal: str,
        }
    
    def get_fingerprint(self) -> str:
        """Generate deterministic fingerprint for deduplication."""
        try:
            id_val = int(self.id)
        except (ValueError, TypeError):
            id_val = None
        return generate_fingerprint(
            broker="XTB",
            tx_type=self.tx_type,
            timestamp=self.date,
            amount=self.amount,
            symbol=self.symbol or "",
            id_value=id_val
        )


def generate_fingerprint(broker: str, tx_type: TransactionType, timestamp: datetime, 
                         amount: Decimal, symbol: str, id_value: Optional[int] = None) -> str:
    """
    Generate a deterministic fingerprint for deduplication.
    
    Fingerprint is based on: broker + type + timestamp + amount + symbol + id
    This ensures each transaction has a unique fingerprint.
    """
    # Normalize timestamp to minute precision for consistency
    ts_normalized = timestamp.replace(second=0, microsecond=0)
    
    id_part = f"{id_value}" if id_value is not None else ""
    fingerprint_data = f"{broker}:{tx_type.value}:{ts_normalized.isoformat()}:{amount}:{symbol}:{id_part}"
    return hashlib.sha256(fingerprint_data.encode()).hexdigest()[:16]


def _safe_decimal(val, default: str = '0') -> Decimal:
    """Convert value to Decimal, returning default for empty/invalid values."""
    if val is None or val == '' or (isinstance(val, str) and val.strip() == ''):
        return Decimal(default)
    try:
        return Decimal(str(val))
    except Exception:
        return Decimal(default)


class _HTMLTableExtractor(HTMLParser):
    """Collect text tables from an HTML document while ignoring scripts/styles."""

    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._table_stack: list[list[list[str]]] = []
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None
        self._capture_cell = False
        self._in_script = False
        self._in_style = False

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag == "script":
            self._in_script = True
            return
        if tag == "style":
            self._in_style = True
            return
        if self._in_script or self._in_style:
            return
        if tag == "table":
            self._table_stack.append([])
        elif tag == "tr" and self._table_stack:
            self._current_row = []
        elif tag in {"td", "th"} and self._current_row is not None:
            self._current_cell = []
            self._capture_cell = True

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "script":
            self._in_script = False
            return
        if tag == "style":
            self._in_style = False
            return
        if self._in_script or self._in_style:
            return
        if tag in {"td", "th"} and self._capture_cell and self._current_row is not None:
            text = " ".join(" ".join(self._current_cell or []).split())
            self._current_row.append(text)
            self._current_cell = None
            self._capture_cell = False
        elif tag == "tr" and self._current_row is not None and self._table_stack:
            if any(cell for cell in self._current_row):
                self._table_stack[-1].append(self._current_row)
            self._current_row = None
        elif tag == "table" and self._table_stack:
            table = self._table_stack.pop()
            if table:
                self.tables.append(table)

    def handle_data(self, data: str) -> None:
        if (
            self._capture_cell
            and not self._in_script
            and not self._in_style
            and data.strip()
        ):
            self._current_cell.append(data.strip())


class XTBParser(ABC):
    """Abstract base class for XTB parsers."""
    
    @abstractmethod
    def parse(self) -> dict:
        """Parse the XTB statement and return extracted data."""
        pass
    
    @abstractmethod
    def normalize_transactions(self, raw_data: dict) -> list[XTBTransaction]:
        """Convert raw parsed data to normalized transaction objects."""
        pass

    def parse_and_normalize(self) -> list[XTBTransaction]:
        """Parse XTB statement and return normalized transactions."""
        return self.normalize_transactions(self.parse())


class XTBExcelParser(XTBParser):
    """Parser for XTB XLSX statement files."""
    
    def __init__(self, file_path: str | Path):
        """
        Initialize XTBExcelParser.
        
        Args:
            file_path: Path to XTB XLSX statement file
        """
        self.file_path = Path(file_path)
        self.workbook = None
        
        if not self.file_path.exists():
            raise XTBFileNotFoundError(f"XTB statement file not found: {self.file_path}")
    
    def parse(self) -> dict:
        """
        Parse XTB XLSX file and extract raw data.
        
        Returns:
            Dictionary with keys: 'open_positions', 'closed_positions', 'cash_operations'
        """
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
        except Exception as e:
            raise XTBFormatError(f"Failed to load XLSX file: {e}")
        
        result = {
            'open_positions': [],
            'closed_positions': [],
            'cash_operations': [],
            'account_info': {}
        }
        
        # Parse each sheet
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            sn_upper = sheet_name.upper()
            
            if 'CASH OPERATION' in sn_upper:
                cash_ops = self._parse_cash_operations(ws)
                result['cash_operations'].extend(cash_ops)
                
            elif 'CLOSED POSITION' in sn_upper:
                closed_pos = self._parse_closed_positions(ws)
                result['closed_positions'].extend(closed_pos)
                
            elif 'OPEN POSITION' in sn_upper:
                open_pos = self._parse_open_positions(ws)
                result['open_positions'].extend(open_pos)
                
            elif 'PENDING ORDERS' in sn_upper:
                pass
        
        self.workbook.close()
        return result
    
    def _find_headers(self, ws, possible_headers: list[str]) -> Optional[int]:
        for row in range(1, ws.max_row + 1):
            col_matches = 0
            expected_headers = ['position', 'symbol', 'type', 'volume', 'id', 'time', 'date']
            for col in range(2, min(ws.max_column + 1, 10)):
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str):
                    val_lower = val.lower().strip()
                    if val_lower in expected_headers or any(h.lower() in val_lower for h in possible_headers):
                        col_matches += 1
            if col_matches >= 3:
                return row
        return None
    
    def _parse_cash_operations(self, ws) -> list[CashOperation]:
        """Parse cash operations from worksheet."""
        operations = []
        
        # Find header row
        header_row = self._find_headers(ws, ['ID', 'Type', 'Time', 'Comment', 'Symbol', 'Amount'])
        
        if header_row is None:
            raise XTBFormatError("Could not find cash operation headers in worksheet")
        
        # Column mapping based on observed format:
        # [None, ID, Type, Time, Comment, Symbol, Amount]
        col_map = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=col).value or "").lower().strip()
            if 'id' in val and val not in ['timestamp', 'time']:
                col_map['id'] = col
            elif val in ['type', 'operation']:
                col_map['type'] = col
            elif val in ['time', 'date', 'timestamp']:
                col_map['time'] = col
            elif 'comment' in val or 'note' in val:
                col_map['comment'] = col
            elif val in ('symbol', 'ticker', 'instrument'):
                # Prefer ticker > symbol > instrument for the symbol column
                if val == 'ticker':
                    col_map['symbol'] = col
                elif val == 'symbol' and col_map.get('symbol') is None:
                    col_map['symbol'] = col
                elif val == 'instrument' and 'symbol' not in col_map:
                    col_map['symbol'] = col
            elif 'amount' in val or val == 'value':
                col_map['amount'] = col
        
        # Validate required columns
        required = ['id', 'type', 'time', 'amount']
        for req in required:
            if req not in col_map:
                raise XTBFormatError(f"Missing required column '{req}' in cash operations")
        
        # Parse data rows
        for row in range(header_row + 1, ws.max_row + 1):
            row_id = ws.cell(row=row, column=col_map['id']).value
            
            # Skip rows without valid ID or non-integer IDs (like "Total" summary rows)
            if row_id is None:
                continue
            try:
                row_id_int = int(row_id)
            except (ValueError, TypeError):
                continue  # Skip non-numeric IDs like "Total"
            
            row_type_raw = ws.cell(row=row, column=col_map['type']).value or ""
            timestamp = self._parse_datetime(ws.cell(row=row, column=col_map['time']).value)
            comment = ws.cell(row=row, column=col_map.get('comment', 1)).value or ""
            symbol = ws.cell(row=row, column=col_map.get('symbol', 1)).value
            amount_str = ws.cell(row=row, column=col_map['amount']).value
            
            if amount_str is None:
                continue
                
            try:
                amount = Decimal(str(amount_str))
            except Exception:
                continue
            
            # Map operation type
            tx_type = self._map_operation_type(str(row_type_raw).lower())
            
            operation = CashOperation(
                operation_id=row_id_int,
                operation_type=tx_type,
                time=timestamp,
                comment=comment if isinstance(comment, str) else str(comment),
                symbol=symbol if isinstance(symbol, str) else str(symbol) if symbol else None,
                amount=amount
            )
            operations.append(operation)
        
        return operations
    
    def _parse_open_positions(self, ws) -> list[OpenPosition]:
        """Parse open positions from worksheet."""
        positions = []
        
        # Find header row
        header_row = self._find_headers(ws, ['Position', 'Symbol', 'Type', 'Volume'])
        
        if header_row is None:
            raise XTBFormatError("Could not find open position headers in worksheet")
        
        # Column mapping based on observed format:
        # [None, Position, Symbol, Type, Volume, Open time, Open price, Market price, 
        #  Purchase value, SL, TP, Margin, Commission, Swap, Rollover, Gross P/L, Comment]
        col_map = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=col).value or "").lower().strip()
            if 'position' in val or val == 'position':
                col_map['position_id'] = col
            elif val in ('symbol', 'ticker', 'instrument'):
                # Prefer ticker > symbol > instrument for the symbol column
                if val == 'ticker':
                    col_map['symbol'] = col
                elif val == 'symbol' and col_map.get('symbol') is None:
                    col_map['symbol'] = col
                elif val == 'instrument' and 'symbol' not in col_map:
                    col_map['symbol'] = col
            elif val == 'type':
                col_map['type'] = col
            elif 'volume' in val or val == 'qty' or val == 'quantity':
                col_map['volume'] = col
            elif 'open time' in val or val == 'opentime':
                col_map['open_time'] = col
            elif 'open price' in val or val == 'entryprice' or val == 'entry price':
                col_map['open_price'] = col
            elif 'market price' in val or val == 'currentprice' or val == 'current price':
                col_map['market_price'] = col
            elif 'purchase' in val or 'value' in val:
                col_map['purchase_value'] = col
            elif val == 'sl' or 'stop' in val:
                col_map['sl'] = col
            elif val == 'tp' or 'target' in val:
                col_map['tp'] = col
            elif 'margin' in val:
                col_map['margin'] = col
            elif 'commission' in val:
                col_map['commission'] = col
            elif 'swap' in val:
                col_map['swap'] = col
            elif 'rollover' in val:
                col_map['rollover'] = col
            elif 'gross' in val or 'p/l' in val or 'pl' in val or 'net' in val:
                col_map['gross_pl'] = col
            elif 'comment' in val:
                col_map['comment'] = col
        
        # Parse data rows
        for row in range(header_row + 1, ws.max_row + 1):
            pos_id = ws.cell(row=row, column=col_map.get('position_id', 1)).value
            if pos_id is None:
                continue
            
            try:
                position_id = int(pos_id)
            except (ValueError, TypeError):
                continue
            
            symbol = ws.cell(row=row, column=col_map.get('symbol', 1)).value
            pos_type_raw = ws.cell(row=row, column=col_map.get('type', 1)).value or ""
            volume_str = ws.cell(row=row, column=col_map.get('volume', 1)).value
            open_time = self._parse_datetime(ws.cell(row=row, column=col_map.get('open_time', 1)).value)
            open_price_str = ws.cell(row=row, column=col_map.get('open_price', 1)).value
            market_price_str = ws.cell(row=row, column=col_map.get('market_price', 1)).value
            purchase_value_str = ws.cell(row=row, column=col_map.get('purchase_value', 1)).value
            sl_str = ws.cell(row=row, column=col_map.get('sl', 1)).value
            tp_str = ws.cell(row=row, column=col_map.get('tp', 1)).value
            margin_str = ws.cell(row=row, column=col_map.get('margin', 1)).value
            commission_str = ws.cell(row=row, column=col_map.get('commission', 1)).value
            swap_str = ws.cell(row=row, column=col_map.get('swap', 1)).value
            rollover_str = ws.cell(row=row, column=col_map.get('rollover', 1)).value
            gross_pl_str = ws.cell(row=row, column=col_map.get('gross_pl', 1)).value
            comment = ws.cell(row=row, column=col_map.get('comment', 1)).value
            
            # Parse datetime fields - skip row if required datetime is missing
            open_time = self._parse_datetime(ws.cell(row=row, column=col_map.get('open_time', 1)).value)
            if open_time is None:
                continue
            
            try:
                volume = Decimal(str(volume_str)) if volume_str else Decimal('0')
                open_price = Decimal(str(open_price_str)) if open_price_str else Decimal('0')
                market_price = Decimal(str(market_price_str)) if market_price_str else None
                purchase_value = Decimal(str(purchase_value_str)) if purchase_value_str else Decimal('0')
                sl = Decimal(str(sl_str)) if sl_str else None
                tp = Decimal(str(tp_str)) if tp_str else None
                margin = Decimal(str(margin_str)) if margin_str else Decimal('0')
                commission = Decimal(str(commission_str)) if commission_str else Decimal('0')
                swap = Decimal(str(swap_str)) if swap_str else Decimal('0')
                rollover = Decimal(str(rollover_str)) if rollover_str else Decimal('0')
                gross_pl = Decimal(str(gross_pl_str)) if gross_pl_str else Decimal('0')
            except Exception:
                continue
            
            position_type = PositionType.BUY if str(pos_type_raw).upper() == 'BUY' else PositionType.SELL
            
            position = OpenPosition(
                position_id=position_id,
                symbol=str(symbol) if symbol else "",
                position_type=position_type,
                volume=volume,
                open_time=open_time,
                open_price=open_price,
                market_price=market_price,
                purchase_value=purchase_value,
                commission=commission,
                swap=swap,
                rollover=rollover,
                gross_pl=gross_pl,
                comment=comment if isinstance(comment, str) else str(comment) if comment else "",
            )
            positions.append(position)
        
        return positions
    
    def _parse_closed_positions(self, ws) -> list[ClosedPosition]:
        positions = []
        
        header_row = self._find_headers(ws, ['Position', 'Symbol', 'Type', 'Volume', 'Open time', 'Close time'])
        
        if header_row is None:
            return positions
        
        col_map = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=col).value or "").lower().strip()
            if 'position id' in val:
                col_map['position_id'] = col
            elif 'position' in val or val == 'instrument':
                col_map['position_id'] = col
            elif val in ('symbol', 'ticker', 'instrument'):
                # Prefer ticker > symbol > instrument for the symbol column
                if val == 'ticker':
                    col_map['symbol'] = col
                elif val == 'symbol' and col_map.get('symbol') is None:
                    col_map['symbol'] = col
                elif val == 'instrument' and 'symbol' not in col_map:
                    col_map['symbol'] = col
            elif val == 'type':
                col_map['type'] = col
            elif 'volume' in val or val == 'qty' or val == 'quantity':
                col_map['volume'] = col
            elif 'open time' in val:
                col_map['open_time'] = col
            elif 'open price' in val or val == 'entryprice' or val == 'entry price':
                col_map['open_price'] = col
            elif 'close time' in val:
                col_map['close_time'] = col
            elif 'close price' in val:
                col_map['close_price'] = col
            elif 'purchase' in val:
                col_map['purchase_value'] = col
            elif 'sale' in val or val == 'sell':
                col_map['sale_value'] = col
            elif val == 'sl' or 'stop' in val:
                col_map['sl'] = col
            elif val == 'tp' or 'target' in val:
                col_map['tp'] = col
            elif 'margin' in val:
                col_map['margin'] = col
            elif 'commission' in val:
                col_map['commission'] = col
            elif 'swap' in val:
                col_map['swap'] = col
            elif 'rollover' in val:
                col_map['rollover'] = col
            elif 'gross' in val or 'p/l' in val or 'pl' in val or 'net' in val:
                col_map['gross_pl'] = col
            elif 'profit' in val:
                col_map['gross_pl'] = col
            elif 'comment' in val:
                col_map['comment'] = col
        
        # Parse data rows
        for row in range(header_row + 1, ws.max_row + 1):
            pos_id = ws.cell(row=row, column=col_map.get('position_id', 1)).value
            if pos_id is None:
                continue
            
            try:
                position_id = int(pos_id)
            except (ValueError, TypeError):
                # If position_id is a string (e.g. instrument name), generate numeric ID from hash
                import hashlib
                position_id = int(hashlib.sha256(str(pos_id).encode()).hexdigest()[:8], 16)
            
            symbol = ws.cell(row=row, column=col_map.get('symbol', 1)).value
            pos_type_raw = ws.cell(row=row, column=col_map.get('type', 1)).value or ""
            volume_str = ws.cell(row=row, column=col_map.get('volume', 1)).value
            open_time = self._parse_datetime(ws.cell(row=row, column=col_map.get('open_time', 1)).value)
            open_price_str = ws.cell(row=row, column=col_map.get('open_price', 1)).value
            close_time = self._parse_datetime(ws.cell(row=row, column=col_map.get('close_time', 1)).value)
            if close_time is None:
                continue
            close_price_str = ws.cell(row=row, column=col_map.get('close_price', 1)).value
            purchase_value_str = ws.cell(row=row, column=col_map.get('purchase_value', 1)).value
            sale_value_str = ws.cell(row=row, column=col_map.get('sale_value', 1)).value
            sl_str = ws.cell(row=row, column=col_map.get('sl', 1)).value
            tp_str = ws.cell(row=row, column=col_map.get('tp', 1)).value
            margin_str = ws.cell(row=row, column=col_map.get('margin', 1)).value
            commission_str = ws.cell(row=row, column=col_map.get('commission', 1)).value
            swap_str = ws.cell(row=row, column=col_map.get('swap', 1)).value
            rollover_str = ws.cell(row=row, column=col_map.get('rollover', 1)).value
            gross_pl_str = ws.cell(row=row, column=col_map.get('gross_pl', 1)).value
            comment = ws.cell(row=row, column=col_map.get('comment', 1)).value
            
            try:
                volume = _safe_decimal(volume_str)
                open_price = _safe_decimal(open_price_str)
                close_price = _safe_decimal(close_price_str)
                purchase_value = _safe_decimal(purchase_value_str)
                sale_value = _safe_decimal(sale_value_str)
                sl = _safe_decimal(sl_str) or None
                tp = _safe_decimal(tp_str) or None
                margin = _safe_decimal(margin_str)
                commission = _safe_decimal(commission_str)
                swap = _safe_decimal(swap_str)
                rollover = _safe_decimal(rollover_str)
                gross_pl = _safe_decimal(gross_pl_str)
            except Exception:
                continue
            
            position_type = PositionType.BUY if str(pos_type_raw).upper() == 'BUY' else PositionType.SELL
            
            position = ClosedPosition(
                position_id=position_id,
                symbol=str(symbol) if symbol else "",
                position_type=position_type,
                volume=volume,
                open_time=open_time,
                open_price=open_price,
                close_time=close_time,
                close_price=close_price,
                purchase_value=purchase_value,
                sale_value=sale_value,
                sl=sl,
                tp=tp,
                margin=margin,
                commission=commission,
                swap=swap,
                rollover=rollover,
                gross_pl=gross_pl,
                comment=comment if isinstance(comment, str) else str(comment) if comment else "",
            )
            positions.append(position)
        
        return positions
    
    def _parse_datetime(self, value) -> datetime:
        """Parse datetime from Excel cell value.
        
        Returns None if value is None or empty string.
        """
        if value is None or value == '' or str(value).strip() == '':
            return None
        
        if isinstance(value, datetime):
            return value
        
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            # Try various formats
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', 
                       '%d/%m/%Y', '%m/%d/%Y %H:%M:%S', '%m/%d/%Y']:
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue
            raise XTBFormatError(f"Could not parse datetime: {value}")
        
        # Excel serial date
        # Use openpyxl's built-in converter for datetime
        from openpyxl.utils import datetime as openpyxl_datetime
        try:
            return openpyxl_datetime.from_timestamp(value)
        except Exception:
            pass
        
        # Manual conversion
        return datetime(*(datetime(1900, 1, 1).toordinal() + int(value) - 2).timetuple()[:6])
    
    def _map_operation_type(self, type_str: str) -> TransactionType:
        """Map operation type string to TransactionType enum."""
        type_lower = type_str.lower()
        
        if type_lower in ['deposit', 'deposit', 'transféré', 'chuyển khoản']:
            return TransactionType.DEPOSIT
        elif type_lower in ['withdrawal', 'withdraw', 'withdrawals', 'transfert']:
            return TransactionType.WITHDRAWAL
        elif type_lower in ['stock purchase', 'stock buy', 'trade', 'position']:
            return TransactionType.OPEN_POSITION
        elif type_lower in ['stock sale', 'stock sell', 'close buy', 'close sell', 'close trade']:
            return TransactionType.CLOSE_POSITION
        elif type_lower in ['dividend', 'div', 'divident']:
            return TransactionType.DIVIDEND
        elif type_lower in ['withholding tax', 'withhold tax', 'tax']:
            return TransactionType.COMMISSION
        elif type_lower in ['commission', 'commission fee']:
            return TransactionType.COMMISSION
        elif type_lower in ['swap', 'swap charge']:
            return TransactionType.SWAP
        elif type_lower in ['stamp duty', 'stamp']:
            return TransactionType.STAMP_DUTY
        else:
            return TransactionType.OTHER
    
    def normalize_transactions(self, raw_data: dict) -> list[XTBTransaction]:
        """
        Convert raw parsed data to normalized transaction objects.
        
        Args:
            raw_data: Dictionary from parse() method
            
        Returns:
            List of XTBTransaction objects
        """
        transactions = []
        
        # Normalize open positions
        for position in raw_data.get('open_positions', []):
            transactions.append(position.to_transaction())
        
        # Normalize closed positions
        for position in raw_data.get('closed_positions', []):
            transactions.append(position.to_transaction())
        
        # Normalize cash operations
        for operation in raw_data.get('cash_operations', []):
            transactions.append(operation.to_transaction())
        
        return transactions
    
    def parse_and_normalize(self) -> list[XTBTransaction]:
        """
        Parse XTB statement and return normalized transactions.
        
        Returns:
            List of XTBTransaction objects
        """
        raw_data = self.parse()
        return self.normalize_transactions(raw_data)


class XTBHTMLParser(XTBExcelParser):
    """Parser for XTB HTML/MHTML statement exports."""

    def parse(self) -> dict:
        tables = self._extract_tables()
        return {
            "open_positions": self._parse_open_position_tables(tables),
            "closed_positions": self._parse_closed_position_tables(tables),
            "cash_operations": self._parse_cash_operation_tables(tables),
            "account_info": {},
        }

    def _extract_tables(self) -> list[list[list[str]]]:
        extractor = _HTMLTableExtractor()
        extractor.feed(self._load_html())
        extractor.close()
        return extractor.tables

    def _load_html(self) -> str:
        raw_bytes = self.file_path.read_bytes()
        suffix = self.file_path.suffix.lower()
        if suffix in {".mhtml", ".mht"}:
            message = BytesParser(policy=policy.default).parsebytes(raw_bytes)
            for part in message.walk():
                if part.get_content_type() == "text/html":
                    return part.get_content()
            raise XTBFormatError("MHTML statement does not contain an HTML body")
        return raw_bytes.decode("utf-8", errors="ignore")

    def _table_contains(self, table: list[list[str]], marker: str) -> bool:
        marker_upper = marker.upper()
        return any(
            marker_upper in " ".join(cell.strip() for cell in row if cell.strip()).upper()
            for row in table
        )

    def _find_header_row(
        self, table: list[list[str]], required_headers: list[str]
    ) -> int | None:
        for index, row in enumerate(table):
            normalized = [cell.lower().strip() for cell in row if cell.strip()]
            if normalized and all(
                any(required in cell for cell in normalized) for required in required_headers
            ):
                return index
        return None

    def _row_value(
        self, row: list[str], col_map: dict[str, int], key: str
    ) -> str | None:
        column = col_map.get(key)
        if column is None or column >= len(row):
            return None
        value = row[column]
        return value if value != "" else None

    def _iter_section_tables(
        self,
        tables: list[list[list[str]]],
        section_marker: str,
        required_headers: list[str],
    ):
        for table in tables:
            if not self._table_contains(table, section_marker):
                continue
            header_index = self._find_header_row(table, required_headers)
            if header_index is None:
                continue
            yield table, header_index

    def _map_open_headers(self, headers: list[str]) -> dict[str, int]:
        col_map: dict[str, int] = {}
        for idx, cell in enumerate(headers):
            val = cell.lower().strip()
            if not val:
                continue
            if "position" in val:
                col_map["position_id"] = idx
            elif val == "symbol":
                col_map["symbol"] = idx
            elif val == "type":
                col_map["type"] = idx
            elif "volume" in val or val in {"qty", "quantity"}:
                col_map["volume"] = idx
            elif "open time" in val or val == "opentime":
                col_map["open_time"] = idx
            elif "open price" in val or val in {"entryprice", "entry price"}:
                col_map["open_price"] = idx
            elif "market price" in val or val in {"currentprice", "current price"}:
                col_map["market_price"] = idx
            elif "purchase" in val:
                col_map["purchase_value"] = idx
            elif val == "sl" or "stop" in val:
                col_map["sl"] = idx
            elif val == "tp" or "target" in val:
                col_map["tp"] = idx
            elif "margin" in val:
                col_map["margin"] = idx
            elif "commission" in val:
                col_map["commission"] = idx
            elif "swap" in val:
                col_map["swap"] = idx
            elif "rollover" in val:
                col_map["rollover"] = idx
            elif "gross" in val or "p/l" in val or val == "pl" or "net" in val:
                col_map["gross_pl"] = idx
            elif "comment" in val:
                col_map["comment"] = idx
        return col_map

    def _map_closed_headers(self, headers: list[str]) -> dict[str, int]:
        col_map = self._map_open_headers(headers)
        for idx, cell in enumerate(headers):
            val = cell.lower().strip()
            if not val:
                continue
            if "close time" in val:
                col_map["close_time"] = idx
            elif "close price" in val:
                col_map["close_price"] = idx
            elif "sale" in val or val == "sell":
                col_map["sale_value"] = idx
        return col_map

    def _map_cash_headers(self, headers: list[str]) -> dict[str, int]:
        col_map: dict[str, int] = {}
        for idx, cell in enumerate(headers):
            val = cell.lower().strip()
            if not val:
                continue
            if "id" in val and val not in {"timestamp", "time"}:
                col_map["id"] = idx
            elif val in {"type", "operation"}:
                col_map["type"] = idx
            elif val in {"time", "date", "timestamp"}:
                col_map["time"] = idx
            elif "comment" in val or "note" in val:
                col_map["comment"] = idx
            elif val in {"symbol", "ticker", "instrument"}:
                col_map["symbol"] = idx
            elif "amount" in val or val == "value":
                col_map["amount"] = idx
        return col_map

    def _parse_open_position_tables(
        self, tables: list[list[list[str]]]
    ) -> list[OpenPosition]:
        positions: list[OpenPosition] = []
        for table, header_index in self._iter_section_tables(
            tables,
            "OPEN POSITION HISTORY",
            ["position", "symbol", "type", "volume", "open time", "open price"],
        ):
            headers = table[header_index]
            col_map = self._map_open_headers(headers)
            for row in table[header_index + 1 :]:
                if row == headers:
                    continue
                first_value = next((cell for cell in row if cell.strip()), "")
                if first_value.lower() == "total":
                    break
                try:
                    position_id = int(self._row_value(row, col_map, "position_id") or "")
                except ValueError:
                    continue
                open_time = self._parse_datetime(self._row_value(row, col_map, "open_time"))
                if open_time is None:
                    continue
                positions.append(
                    OpenPosition(
                        position_id=position_id,
                        symbol=str(self._row_value(row, col_map, "symbol") or ""),
                        position_type=PositionType.BUY
                        if str(self._row_value(row, col_map, "type") or "").upper() == "BUY"
                        else PositionType.SELL,
                        volume=_safe_decimal(self._row_value(row, col_map, "volume")),
                        open_time=open_time,
                        open_price=_safe_decimal(self._row_value(row, col_map, "open_price")),
                        market_price=_safe_decimal(
                            self._row_value(row, col_map, "market_price"), default="0"
                        )
                        if self._row_value(row, col_map, "market_price") is not None
                        else None,
                        purchase_value=_safe_decimal(
                            self._row_value(row, col_map, "purchase_value")
                        ),
                        commission=_safe_decimal(
                            self._row_value(row, col_map, "commission")
                        ),
                        swap=_safe_decimal(self._row_value(row, col_map, "swap")),
                        rollover=_safe_decimal(
                            self._row_value(row, col_map, "rollover")
                        ),
                        gross_pl=_safe_decimal(self._row_value(row, col_map, "gross_pl")),
                        comment=str(self._row_value(row, col_map, "comment") or ""),
                    )
                )
        return positions

    def _parse_closed_position_tables(
        self, tables: list[list[list[str]]]
    ) -> list[ClosedPosition]:
        positions: list[ClosedPosition] = []
        for table, header_index in self._iter_section_tables(
            tables,
            "CLOSED POSITION HISTORY",
            ["position", "symbol", "type", "volume", "open time", "close time"],
        ):
            headers = table[header_index]
            col_map = self._map_closed_headers(headers)
            for row in table[header_index + 1 :]:
                if row == headers:
                    continue
                first_value = next((cell for cell in row if cell.strip()), "")
                if first_value.lower() == "total":
                    break
                try:
                    position_id = int(self._row_value(row, col_map, "position_id") or "")
                except ValueError:
                    continue
                close_time = self._parse_datetime(
                    self._row_value(row, col_map, "close_time")
                )
                if close_time is None:
                    continue
                positions.append(
                    ClosedPosition(
                        position_id=position_id,
                        symbol=str(self._row_value(row, col_map, "symbol") or ""),
                        position_type=PositionType.BUY
                        if str(self._row_value(row, col_map, "type") or "").upper() == "BUY"
                        else PositionType.SELL,
                        volume=_safe_decimal(self._row_value(row, col_map, "volume")),
                        open_time=self._parse_datetime(
                            self._row_value(row, col_map, "open_time")
                        ),
                        open_price=_safe_decimal(self._row_value(row, col_map, "open_price")),
                        close_time=close_time,
                        close_price=_safe_decimal(
                            self._row_value(row, col_map, "close_price")
                        ),
                        purchase_value=_safe_decimal(
                            self._row_value(row, col_map, "purchase_value")
                        ),
                        sale_value=_safe_decimal(
                            self._row_value(row, col_map, "sale_value")
                        ),
                        sl=_safe_decimal(self._row_value(row, col_map, "sl")) or None,
                        tp=_safe_decimal(self._row_value(row, col_map, "tp")) or None,
                        margin=_safe_decimal(self._row_value(row, col_map, "margin")),
                        commission=_safe_decimal(
                            self._row_value(row, col_map, "commission")
                        ),
                        swap=_safe_decimal(self._row_value(row, col_map, "swap")),
                        rollover=_safe_decimal(
                            self._row_value(row, col_map, "rollover")
                        ),
                        gross_pl=_safe_decimal(self._row_value(row, col_map, "gross_pl")),
                        comment=str(self._row_value(row, col_map, "comment") or ""),
                    )
                )
        return positions

    def _parse_cash_operation_tables(
        self, tables: list[list[list[str]]]
    ) -> list[CashOperation]:
        operations: list[CashOperation] = []
        for table, header_index in self._iter_section_tables(
            tables,
            "CASH OPERATION HISTORY",
            ["id", "type", "time", "amount"],
        ):
            headers = table[header_index]
            col_map = self._map_cash_headers(headers)
            for row in table[header_index + 1 :]:
                if row == headers:
                    continue
                first_value = next((cell for cell in row if cell.strip()), "")
                if first_value.lower() == "total":
                    break
                try:
                    operation_id = int(self._row_value(row, col_map, "id") or "")
                except ValueError:
                    continue
                timestamp = self._parse_datetime(self._row_value(row, col_map, "time"))
                if timestamp is None:
                    continue
                amount = self._row_value(row, col_map, "amount")
                if amount is None:
                    continue
                operations.append(
                    CashOperation(
                        operation_id=operation_id,
                        operation_type=self._map_operation_type(
                            str(self._row_value(row, col_map, "type") or "").lower()
                        ),
                        time=timestamp,
                        comment=str(self._row_value(row, col_map, "comment") or ""),
                        symbol=self._row_value(row, col_map, "symbol"),
                        amount=_safe_decimal(amount),
                    )
                )
        return operations


class XTBDailyStatementPdfParser(XTBParser):
    """Parser for XTB daily executed-order PDF statements."""

    _ROW_RE = re.compile(
        r"^\s*\d+\s+"
        r"(?P<order_id>\d+)\s+"
        r"(?P<symbol>[A-Z0-9.]+)\s+"
        r"(?P<instrument_head>.*?)\s+"
        r"(?P<venue>[A-Z]{3,5})\s+"
        r"(?P<quantity>\d+[.,]\d+)\s+"
        r"(?P<date>\d{2}/\d{2}/\d{4})\s+"
        r"(?P<order_mode>[A-Za-z]+)\s+"
        r"(?P<execution_price>\d+[.,]\d+)\s+"
        r"(?P<total_value>\d+[.,]\d+)\s+"
        r"(?P<asset_type>.*?)\s+"
        r"(?P<currency>[A-Z]{3})\s+"
        r"(?P<fx_rate>\d+[.,]\d+)\s+"
        r"(?P<conversion_fee>\d+[.,]\d+)\s+"
        r"(?P<commission>\d+[.,]\d+)\s+"
        r"(?P<total_cost>\d+[.,]\d+)\s*$"
    )
    _CONTINUATION_RE = re.compile(
        r"^\s*(?P<instrument_tail>.*?)\s+"
        r"(?P<time>\d{2}:\d{2}:\d{2})\s*$"
    )

    def __init__(
        self,
        file_path: str | Path | None = None,
        *,
        password: str | None = None,
        layout_text: str | None = None,
    ):
        self.file_path = Path(file_path) if file_path is not None else None
        self.password = password
        self._layout_text = layout_text

        if self.file_path is not None and not self.file_path.exists():
            raise XTBFileNotFoundError(
                f"XTB statement file not found: {self.file_path}"
            )

    @classmethod
    def from_layout_text(cls, layout_text: str) -> "XTBDailyStatementPdfParser":
        """Build a parser from already-extracted pdftotext -layout output."""
        return cls(layout_text=layout_text)

    def parse(self) -> dict:
        text = (
            self._layout_text
            if self._layout_text is not None
            else self._extract_text()
        )
        return {
            "daily_trades": self._parse_daily_trades(text),
            "account_info": {},
        }

    def normalize_transactions(self, raw_data: dict) -> list[XTBTransaction]:
        return [
            trade.to_transaction()
            for trade in raw_data.get("daily_trades", [])
        ]

    def _extract_text(self) -> str:
        if self.file_path is None:
            raise XTBFormatError("PDF parser requires a file path or layout text")

        command = ["pdftotext", "-layout"]
        if self.password:
            command.extend(["-upw", self.password])
        command.extend([str(self.file_path), "-"])

        try:
            completed = subprocess.run(  # noqa: S603
                command,
                check=False,
                capture_output=True,
                text=True,
                timeout=30,
            )
        except FileNotFoundError as exc:
            raise XTBFormatError(
                "pdftotext is required to parse XTB PDF statements"
            ) from exc
        except subprocess.TimeoutExpired as exc:
            raise XTBFormatError(
                "Timed out extracting text from XTB PDF statement"
            ) from exc

        if completed.returncode != 0:
            stderr = completed.stderr.strip()
            if "incorrect password" in stderr.lower() or "password" in stderr.lower():
                raise XTBFormatError(
                    "XTB PDF is encrypted and requires a valid password"
                )
            raise XTBFormatError("Failed to extract text from XTB PDF statement")

        return completed.stdout

    def _parse_daily_trades(self, text: str) -> list[XTBDailyStatementTrade]:
        trades: list[XTBDailyStatementTrade] = []
        current_position_type = PositionType.BUY
        lines = text.splitlines()
        index = 0

        while index < len(lines):
            line = lines[index]
            normalized_line = _strip_accents(line).lower()
            if "lenh ban" in normalized_line:
                current_position_type = PositionType.SELL
            elif "lenh mua" in normalized_line:
                current_position_type = PositionType.BUY

            match = self._ROW_RE.match(line)
            if match is None:
                index += 1
                continue

            continuation = self._find_continuation(lines, index + 1)
            if continuation is None:
                raise XTBFormatError("XTB PDF trade row is missing execution time")

            instrument_tail, time_text = continuation
            date_text = match.group("date")
            trade_time = datetime.strptime(
                f"{date_text} {time_text}", "%d/%m/%Y %H:%M:%S"
            )
            trades.append(
                XTBDailyStatementTrade(
                    order_id=int(match.group("order_id")),
                    symbol=match.group("symbol"),
                    instrument_name=self._join_instrument_name(
                        match.group("instrument_head"),
                        instrument_tail,
                    ),
                    position_type=current_position_type,
                    quantity=_parse_pdf_decimal(match.group("quantity")),
                    trade_time=trade_time,
                    execution_price=_parse_pdf_decimal(match.group("execution_price")),
                    total_value=_parse_pdf_decimal(match.group("total_value")),
                    asset_type=" ".join(match.group("asset_type").split()),
                    currency=match.group("currency"),
                    fx_rate=_parse_pdf_decimal(match.group("fx_rate")),
                    conversion_fee=_parse_pdf_decimal(match.group("conversion_fee")),
                    commission=_parse_pdf_decimal(match.group("commission")),
                    total_cost=_parse_pdf_decimal(match.group("total_cost")),
                )
            )
            index += 1

        if not trades:
            raise XTBFormatError(
                "No executed trade rows found in XTB daily PDF statement"
            )

        return trades

    def _find_continuation(
        self, lines: list[str], start_index: int
    ) -> tuple[str, str] | None:
        for line in lines[start_index : min(start_index + 4, len(lines))]:
            match = self._CONTINUATION_RE.match(line)
            if match is not None:
                return (
                    " ".join(match.group("instrument_tail").split()),
                    match.group("time"),
                )
        return None

    def _join_instrument_name(self, head: str, tail: str) -> str:
        normalized_head = " ".join(head.split())
        if not tail:
            return normalized_head.rstrip(",")
        separator = ", " if normalized_head.endswith(",") else " "
        return f"{normalized_head.rstrip(',')}{separator}{tail}"


def _parse_pdf_decimal(value: str) -> Decimal:
    """Parse XTB PDF numeric values that use comma decimal separators."""
    normalized = value.strip()
    if "," in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    return Decimal(normalized)


def _strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def parse_xtb_statement(
    file_path: str | Path, *, pdf_password: str | None = None
) -> list[XTBTransaction]:
    """
    Convenience function to parse XTB statement and return normalized transactions.
    
    Args:
        file_path: Path to XTB XLSX, HTML/MHTML, or daily PDF statement file
        
    Returns:
        List of XTBTransaction objects
        
    Raises:
        XTBFileNotFoundError: If file doesn't exist
        XTBFormatError: If file format is invalid
    """
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in {".html", ".mhtml", ".mht"}:
        parser: XTBParser = XTBHTMLParser(path)
    elif suffix == ".pdf":
        parser = XTBDailyStatementPdfParser(path, password=pdf_password)
    else:
        parser = XTBExcelParser(path)
    return parser.parse_and_normalize()
