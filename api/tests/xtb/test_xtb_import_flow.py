from __future__ import annotations

import os
from datetime import datetime
from io import BytesIO

import pytest
import pytest_asyncio
from app.config import settings
from app.db.base import Base
from app.db.models import ImportArtifact, Transaction
from app.db.safety import (
    DEFAULT_TEST_DATABASE_SERVER_URL,
    build_temporary_test_database_url,
    pick_safe_test_database_server_url,
    quote_postgresql_identifier,
)
from app.services.xtb_ingest import (
    _repair_legacy_xtb_split_transactions,
    confirm_import,
    parse_xtb_file,
)
from openpyxl import Workbook
from sqlalchemy import select, text
from sqlalchemy.engine import make_url
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from sqlalchemy.pool import NullPool

SYNTHETIC_XTB_ACCOUNT = "XTB-SYNTHETIC-ACCOUNT"


def _resolve_database_url():
    return make_url(
        pick_safe_test_database_server_url(
            os.environ.get("TEST_DATABASE_BASE_URL")
            or os.environ.get("DATABASE_URL")
            or settings.DATABASE_URL,
            default_url=DEFAULT_TEST_DATABASE_SERVER_URL,
        )
    )


async def _create_temporary_database() -> str:
    base_url = _resolve_database_url()
    if base_url.get_backend_name() != "postgresql":
        pytest.skip("xtb import flow tests require PostgreSQL")

    database_url = build_temporary_test_database_url(
        base_url,
        name_prefix="portfolio_tracker_xtb",
        context="api/tests/xtb/test_xtb_import_flow.py",
    )
    temp_url = make_url(database_url)
    quoted_database_name = quote_postgresql_identifier(temp_url.database or "")
    admin_url = base_url.set(database="postgres").render_as_string(hide_password=False)
    admin_engine = create_async_engine(
        admin_url,
        isolation_level="AUTOCOMMIT",
        pool_pre_ping=True,
        poolclass=NullPool,
    )
    try:
        async with admin_engine.connect() as conn:
            await conn.execute(text(f"CREATE DATABASE {quoted_database_name}"))
    except Exception as exc:
        await admin_engine.dispose()
        pytest.skip(f"unable to create disposable postgres database: {exc}")

    await admin_engine.dispose()
    return database_url


async def _drop_temporary_database(database_url: str) -> None:
    temp_url = make_url(database_url)
    quoted_database_name = quote_postgresql_identifier(temp_url.database or "")
    admin_url = temp_url.set(database="postgres").render_as_string(hide_password=False)
    admin_engine = create_async_engine(
        admin_url,
        isolation_level="AUTOCOMMIT",
        pool_pre_ping=True,
        poolclass=NullPool,
    )
    try:
        async with admin_engine.connect() as conn:
            await conn.execute(
                text(
                    "SELECT pg_terminate_backend(pid) "
                    "FROM pg_stat_activity "
                    "WHERE datname = :database_name AND pid <> pg_backend_pid()"
                ),
                {"database_name": temp_url.database},
            )
            await conn.execute(text(f"DROP DATABASE IF EXISTS {quoted_database_name}"))
    finally:
        await admin_engine.dispose()


@pytest_asyncio.fixture
async def test_session_factory():
    database_url = await _create_temporary_database()
    try:
        engine = create_async_engine(
            database_url,
            pool_pre_ping=True,
            poolclass=NullPool,
        )
        session_factory = async_sessionmaker(engine, expire_on_commit=False)

        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        try:
            yield session_factory
        finally:
            await engine.dispose()
    finally:
        await _drop_temporary_database(database_url)


def _build_latest_style_xtb_workbook() -> bytes:
    wb = Workbook()
    ws_closed = wb.active
    ws_closed.title = "Closed Positions"
    ws_closed.append(["Account", SYNTHETIC_XTB_ACCOUNT])
    ws_closed.append(["Closed Positions", ""])
    ws_closed.append(["Date from (UTC)", datetime(2026, 4, 1, 0, 0)])
    ws_closed.append(["Date to (UTC)", datetime(2026, 4, 17, 12, 28, 56)])
    ws_closed.append(
        [
            "Instrument",
            "Category",
            "Ticker",
            "Type",
            "Volume",
            "Open Price",
            "Open Time (UTC)",
            "Close Price",
            "Close Time (UTC)",
            "Product",
            "Profit/Loss",
            "Gross Profit",
            "Purchase Value",
            "Sale Value",
            "Stop Loss",
            "Take Profit",
            "Commission",
            "Margin",
            "Swap",
            "Rollover",
            "Open Conversion Rate",
            "Close Conversion Rate",
            "Close Origin",
            "Position ID",
            "Comment",
        ]
    )

    ws_cash = wb.create_sheet("Cash Operations")
    ws_cash.append(["Account number", SYNTHETIC_XTB_ACCOUNT])
    ws_cash.append(["Cash Operations", ""])
    ws_cash.append(["Date from (UTC)", datetime(2026, 4, 1, 0, 0)])
    ws_cash.append(["Date to (UTC)", datetime(2026, 4, 17, 12, 28, 56)])
    ws_cash.append(
        ["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment", "Product"]
    )
    ws_cash.append(
        [
            "Stock purchase",
            "SOI.FR",
            "Soitec",
            datetime(2026, 4, 16, 9, 0),
            -100.0,
            "1001",
            "OPEN BUY 1 SOI.FR",
            "My Trades",
        ]
    )
    ws_cash.append(
        [
            "Dividend",
            "SOI.FR",
            "Soitec",
            datetime(2026, 4, 16, 9, 5),
            2.5,
            "1002",
            "cash dividend",
            "My Trades",
        ]
    )
    ws_cash.append(
        [
            "Dividend",
            "SOI.FR",
            "Soitec",
            datetime(2026, 4, 16, 9, 6),
            -2.5,
            "1003",
            "corr SOI.FR dividend adjustment",
            "My Trades",
        ]
    )

    buff = BytesIO()
    wb.save(buff)
    return buff.getvalue()


def _build_workbook_with_duplicate_close_position() -> bytes:
    wb = Workbook()
    ws_closed = wb.active
    ws_closed.title = "Closed Positions"
    ws_closed.append(["Account", SYNTHETIC_XTB_ACCOUNT])
    ws_closed.append(["Closed Positions", ""])
    ws_closed.append(["Date from (UTC)", datetime(2026, 4, 1, 0, 0)])
    ws_closed.append(["Date to (UTC)", datetime(2026, 4, 17, 12, 28, 56)])
    ws_closed.append(
        [
            "Instrument",
            "Category",
            "Ticker",
            "Type",
            "Volume",
            "Open Price",
            "Open Time (UTC)",
            "Close Price",
            "Close Time (UTC)",
            "Product",
            "Profit/Loss",
            "Gross Profit",
            "Purchase Value",
            "Sale Value",
            "Stop Loss",
            "Take Profit",
            "Commission",
            "Margin",
            "Swap",
            "Rollover",
            "Open Conversion Rate",
            "Close Conversion Rate",
            "Close Origin",
            "Position ID",
            "Comment",
        ]
    )
    ws_closed.append(
        [
            "Soitec",
            "STOCK",
            "SOI.FR",
            "BUY",
            1.0,
            100.0,
            datetime(2026, 4, 16, 9, 0),
            110.0,
            datetime(2026, 4, 17, 9, 0),
            "My Trades",
            10.0,
            10.0,
            100.0,
            110.0,
            None,
            None,
            0.0,
            0.0,
            0.0,
            0.0,
            1.0,
            1.0,
            "cash",
            "2001",
            "CLOSE BUY 1 SOI.FR",
        ]
    )

    ws_cash = wb.create_sheet("Cash Operations")
    ws_cash.append(["Account number", SYNTHETIC_XTB_ACCOUNT])
    ws_cash.append(["Cash Operations", ""])
    ws_cash.append(["Date from (UTC)", datetime(2026, 4, 1, 0, 0)])
    ws_cash.append(["Date to (UTC)", datetime(2026, 4, 17, 12, 28, 56)])
    ws_cash.append(
        ["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment", "Product"]
    )
    ws_cash.append(
        [
            "Stock purchase",
            "SOI.FR",
            "Soitec",
            datetime(2026, 4, 16, 9, 0),
            -100.0,
            "1001",
            "OPEN BUY 1 SOI.FR",
            "My Trades",
        ]
    )
    ws_cash.append(
        [
            "Stock sell",
            "SOI.FR",
            "Soitec",
            datetime(2026, 4, 17, 9, 0),
            110.0,
            "1002",
            "CLOSE BUY 1 @ 110.0",
            "My Trades",
        ]
    )

    buff = BytesIO()
    wb.save(buff)
    return buff.getvalue()


@pytest.mark.asyncio
async def test_parse_and_confirm_xtb_import_skips_corrections_and_dedupes_reimports(
    test_session_factory,
):
    file_bytes = _build_latest_style_xtb_workbook()

    async with test_session_factory() as session:
        artifact = await parse_xtb_file(file_bytes, "latest_style.xlsx", session)
        assert artifact.status == "reviewed"
        assert artifact.parsed_count == 3
        assert artifact.duplicate_count == 1

        result = await confirm_import(artifact.id, session)
        assert result["committed"] == 2
        assert result["duplicates_skipped"] == 1

    async with test_session_factory() as session:
        txs = (
            (await session.execute(select(Transaction).order_by(Transaction.timestamp)))
            .scalars()
            .all()
        )
        assert len(txs) == 2
        assert {tx.asset_symbol for tx in txs} == {"SOI.FR", "USD"}
        assert {tx.tx_type for tx in txs} == {"dividend", "buy"}

    async with test_session_factory() as session:
        artifact_2 = await parse_xtb_file(file_bytes, "latest_style.xlsx", session)
        assert artifact_2.status == "reviewed"
        assert artifact_2.parsed_count == 3
        assert artifact_2.duplicate_count == 3

        result_2 = await confirm_import(artifact_2.id, session)
        assert result_2["committed"] == 0
        assert result_2["duplicates_skipped"] == 3

        artifacts = (
            (await session.execute(select(ImportArtifact).order_by(ImportArtifact.id)))
            .scalars()
            .all()
        )
        assert [artifact.committed_count for artifact in artifacts] == [2, 0]


@pytest.mark.asyncio
async def test_parse_and_confirm_xtb_import_ignores_closed_position_rows_when_cash_sell_exists(
    test_session_factory,
):
    file_bytes = _build_workbook_with_duplicate_close_position()

    async with test_session_factory() as session:
        artifact = await parse_xtb_file(file_bytes, "duplicate_close.xlsx", session)
        result = await confirm_import(artifact.id, session)

        assert artifact.parsed_count == 3
        assert result["committed"] == 2
        assert result["duplicates_skipped"] == 1

    async with test_session_factory() as session:
        txs = (
            (await session.execute(select(Transaction).order_by(Transaction.timestamp)))
            .scalars()
            .all()
        )
        assert len(txs) == 2
        assert [(tx.tx_type, tx.asset_symbol, str(tx.quantity)) for tx in txs] == [
            ("buy", "SOI.FR", "1.0000000000"),
            ("sell", "SOI.FR", "1.0000000000"),
        ]


@pytest.mark.asyncio
async def test_legacy_split_repair_skips_correction_rows(test_session_factory):
    async with test_session_factory() as session:
        session.add_all(
            [
                Transaction(
                    institution="xtb",
                    tx_type="buy",
                    asset_symbol="XLU.US",
                    asset_type="equity",
                    quantity=1,
                    price_usd=1,
                    total_usd=1,
                    fee=0,
                    fee_currency="USD",
                    timestamp=datetime(2025, 9, 25, 15, 30, 1),
                    fingerprint="corr-split-row",
                    raw_data={"description": "corr XLU.US split 2 for 1"},
                ),
                Transaction(
                    institution="xtb",
                    tx_type="buy",
                    asset_symbol="XLU.US",
                    asset_type="equity",
                    quantity=6,
                    price_usd=86.17,
                    total_usd=517.02,
                    fee=0,
                    fee_currency="USD",
                    timestamp=datetime(2025, 9, 25, 15, 30, 2),
                    fingerprint="real-split-row",
                    raw_data={"description": "XLU.US split 2 for 1"},
                ),
            ]
        )
        await session.commit()

        repaired = await _repair_legacy_xtb_split_transactions(session)
        await session.commit()

        rows = (
            (
                await session.execute(
                    select(Transaction).order_by(
                        Transaction.timestamp.asc(), Transaction.id.asc()
                    )
                )
            )
            .scalars()
            .all()
        )

        assert repaired == 1
        assert len(rows) == 2
        assert rows[0].raw_data["description"] == "corr XLU.US split 2 for 1"
        assert rows[0].tx_type == "buy"
        assert rows[1].tx_type == "split"
        assert str(rows[1].quantity) == "2.0000000000"


@pytest.mark.asyncio
async def test_legacy_split_repair_raises_on_unsupported_split_descriptions(
    test_session_factory,
):
    async with test_session_factory() as session:
        session.add(
            Transaction(
                institution="xtb",
                tx_type="buy",
                asset_symbol="XLU.US",
                asset_type="equity",
                quantity=1,
                price_usd=1,
                total_usd=1,
                fee=0,
                fee_currency="USD",
                timestamp=datetime(2025, 9, 25, 15, 30, 1),
                fingerprint="bad-legacy-split-row",
                raw_data={"description": "XLU.US split soon"},
            )
        )
        await session.commit()

        with pytest.raises(
            ValueError, match="Unsupported legacy XTB split description"
        ):
            await _repair_legacy_xtb_split_transactions(session)
